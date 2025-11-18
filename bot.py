import logging
import gspread
import pandas as pd
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
from datetime import datetime, timedelta
import os
import time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Токен вашего бота
BOT_TOKEN = "7850122522:AAEZD921qLeR24BKaSqoGNgCdA6GAaBsf9I"

# Глобальные переменные
user_data = {}
spreadsheet = None
managers_sheet = None
employees_sheet = None
time_sheet = None

def initialize_google_sheets():
    """Инициализация подключения к Google Sheets"""
    global spreadsheet, managers_sheet, employees_sheet, time_sheet
    
    try:
        # Проверяем наличие файла credentials
        if not os.path.exists('credentials.json'):
            print("❌ Файл credentials.json не найден!")
            print("📁 Создайте его по инструкции и поместите в папку с ботом")
            return False
        
        # Подключаемся к Google Sheets
        gc = gspread.service_account(filename='credentials.json')
        print("✅ Успешное подключение к Google Sheets!")
        
        # Пробуем открыть таблицу
        try:
            spreadsheet = gc.open("WorkTimeTracker")
            print("✅ Таблица 'WorkTimeTracker' найдена!")
        except gspread.SpreadsheetNotFound:
            print("❌ Таблица 'WorkTimeTracker' не найдена!")
            print("📝 Создайте таблицу с именем 'WorkTimeTracker' в Google Sheets")
            return False
        
        # Получаем все существующие листы
        existing_sheets = [sheet.title for sheet in spreadsheet.worksheets()]
        print(f"📊 Существующие листы: {existing_sheets}")
        
        # Создаем или получаем листы
        managers_sheet = get_or_create_worksheet("Менеджеры", [['TelegramID', 'ФИО_менеджера']], existing_sheets)
        employees_sheet = get_or_create_worksheet("Сотрудники", [['ID', 'ФИО', 'TelegramID', 'Активен']], existing_sheets)
        time_sheet = get_or_create_worksheet("РабочееВремя", [['ID', 'ID_сотрудника', 'Дата', 'Часы', 'Тип', 'Дата_ввода']], existing_sheets)
        
        print("✅ Все листы инициализированы успешно!")
        return True
        
    except Exception as e:
        print(f"❌ Ошибка инициализации Google Sheets: {e}")
        return False

def get_or_create_worksheet(name, headers, existing_sheets=None):
    """Создает лист если его нет, или возвращает существующий"""
    try:
        # Если не передали список существующих листов, получаем его
        if existing_sheets is None:
            existing_sheets = [sheet.title for sheet in spreadsheet.worksheets()]
        
        # Проверяем существует ли лист
        if name in existing_sheets:
            worksheet = spreadsheet.worksheet(name)
            print(f"✅ Лист '{name}' найден")
            return worksheet
        else:
            print(f"📝 Создаю лист '{name}'...")
            worksheet = spreadsheet.add_worksheet(title=name, rows=100, cols=len(headers[0]))
            worksheet.update('A1', headers)
            print(f"✅ Лист '{name}' создан")
            return worksheet
    except Exception as e:
        print(f"❌ Ошибка при работе с листом '{name}': {e}")
        # Пробуем получить лист, если он уже существует
        try:
            worksheet = spreadsheet.worksheet(name)
            print(f"✅ Лист '{name}' найден после ошибки")
            return worksheet
        except:
            raise

def is_manager(user_id):
    """Проверяет является ли пользователь менеджером"""
    try:
        managers = managers_sheet.get_all_records()
        for manager in managers:
            if str(manager['TelegramID']) == str(user_id):
                return True
        return False
    except Exception as e:
        logger.error(f"Ошибка проверки менеджера: {e}")
        return False

def create_calendar(year, month):
    """Создает календарь для выбора даты"""
    keyboard = []
    
    # Заголовок с месяцем и годом
    month_name = datetime(year, month, 1).strftime('%B %Y')
    keyboard.append([InlineKeyboardButton(month_name, callback_data="ignore")])
    
    # Дни недели
    days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    keyboard.append([InlineKeyboardButton(day, callback_data="ignore") for day in days])
    
    # Числа месяца
    first_day = datetime(year, month, 1)
    last_day = datetime(year, month + 1, 1) - timedelta(days=1) if month < 12 else datetime(year + 1, 1, 1) - timedelta(days=1)
    
    week = []
    # Пустые клетки до первого дня
    for _ in range(first_day.weekday()):
        week.append(InlineKeyboardButton(" ", callback_data="ignore"))
    
    # Числа
    for day in range(1, last_day.day + 1):
        week.append(InlineKeyboardButton(str(day), callback_data=f"calendar_{year}_{month}_{day}"))
        if len(week) == 7:
            keyboard.append(week)
            week = []
    
    if week:
        keyboard.append(week)
    
    # Кнопки навигации
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    nav_buttons = [
        InlineKeyboardButton("◀️", callback_data=f"nav_{prev_year}_{prev_month}"),
        InlineKeyboardButton("▶️", callback_data=f"nav_{next_year}_{next_month}")
    ]
    keyboard.append(nav_buttons)
    
    return keyboard

def generate_excel_report(employee_data, period_start=None, period_end=None):
    """Генерирует Excel отчет по данным сотрудников"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по рабочему времени"
    
    # Заголовок
    period_text = ""
    if period_start and period_end:
        period_text = f" за период с {period_start} по {period_end}"
    
    ws['A1'] = f"Отчет по рабочему времени{period_text}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Заголовки таблицы
    headers = ['ФИО сотрудника', 'Дата', 'Дневные часы', 'Ночные часы', 'Всего за день', 'Примечание']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Данные
    row = 4
    for employee_name, dates in employee_data.items():
        # Строка с именем сотрудника
        ws.cell(row=row, column=1, value=employee_name).font = Font(bold=True)
        row += 1
        
        total_day_hours = 0
        total_night_hours = 0
        days_with_data = 0
        
        # Создаем полный список дат в периоде
        if period_start and period_end:
            try:
                start_date = datetime.strptime(period_start, '%d.%m.%Y')
                end_date = datetime.strptime(period_end, '%d.%m.%Y')
                all_dates = []
                current_date = start_date
                while current_date <= end_date:
                    all_dates.append(current_date.strftime('%d.%m.%Y'))
                    current_date += timedelta(days=1)
                
                # Добавляем все даты в отчет
                for date_str in all_dates:
                    shifts = dates.get(date_str, {'день': 0, 'ночь': 0})
                    day_hours = shifts.get('день', 0)
                    night_hours = shifts.get('ночь', 0)
                    
                    # Проверяем, есть ли вообще данные за этот день
                    has_day_data = day_hours != 0
                    has_night_data = night_hours != 0
                    has_any_data = has_day_data or has_night_data
                    
                    if has_any_data:
                        total_day = day_hours + night_hours
                        total_day_hours += day_hours
                        total_night_hours += night_hours
                        days_with_data += 1
                        
                        # Форматируем данные для Excel
                        day_display = day_hours if has_day_data else "-"
                        night_display = night_hours if has_night_data else "-"
                        total_display = total_day
                    else:
                        # Если нет данных за день - прочерки
                        day_display = "-"
                        night_display = "-"
                        total_display = "-"
                    
                    ws.cell(row=row, column=2, value=date_str)
                    ws.cell(row=row, column=3, value=day_display)
                    ws.cell(row=row, column=4, value=night_display)
                    ws.cell(row=row, column=5, value=total_display)
                    
                    # Примечание о превышении 24 часов (только если есть данные)
                    if has_any_data and total_day > 24:
                        ws.cell(row=row, column=6, value="⚠️ Превышено 24 часа")
                    
                    row += 1
            except ValueError as e:
                print(f"Ошибка при создании списка дат: {e}")
                # Если ошибка - используем старый метод (только даты с данными)
                for date, shifts in sorted(dates.items()):
                    day_hours = shifts.get('день', 0)
                    night_hours = shifts.get('ночь', 0)
                    
                    # Проверяем, есть ли вообще данные за этот день
                    has_day_data = day_hours != 0
                    has_night_data = night_hours != 0
                    has_any_data = has_day_data or has_night_data
                    
                    if has_any_data:
                        total_day = day_hours + night_hours
                        total_day_hours += day_hours
                        total_night_hours += night_hours
                        days_with_data += 1
                        
                        # Форматируем данные для Excel
                        day_display = day_hours if has_day_data else "-"
                        night_display = night_hours if has_night_data else "-"
                        total_display = total_day
                    else:
                        # Если нет данных за день - прочерки
                        day_display = "-"
                        night_display = "-"
                        total_display = "-"
                    
                    ws.cell(row=row, column=2, value=date)
                    ws.cell(row=row, column=3, value=day_display)
                    ws.cell(row=row, column=4, value=night_display)
                    ws.cell(row=row, column=5, value=total_display)
                    
                    # Примечание о превышении 24 часов (только если есть данные)
                    if has_any_data and total_day > 24:
                        ws.cell(row=row, column=6, value="⚠️ Превышено 24 часа")
                    
                    row += 1
        else:
            # Если период не указан, используем только даты с данными
            for date, shifts in sorted(dates.items()):
                day_hours = shifts.get('день', 0)
                night_hours = shifts.get('ночь', 0)
                
                # Проверяем, есть ли вообще данные за этот день
                has_day_data = day_hours != 0
                has_night_data = night_hours != 0
                has_any_data = has_day_data or has_night_data
                
                if has_any_data:
                    total_day = day_hours + night_hours
                    total_day_hours += day_hours
                    total_night_hours += night_hours
                    days_with_data += 1
                    
                    # Форматируем данные для Excel
                    day_display = day_hours if has_day_data else "-"
                    night_display = night_hours if has_night_data else "-"
                    total_display = total_day
                else:
                    # Если нет данных за день - прочерки
                    day_display = "-"
                    night_display = "-"
                    total_display = "-"
                
                ws.cell(row=row, column=2, value=date)
                ws.cell(row=row, column=3, value=day_display)
                ws.cell(row=row, column=4, value=night_display)
                ws.cell(row=row, column=5, value=total_display)
                
                # Примечание о превышении 24 часов (только если есть данные)
                if has_any_data and total_day > 24:
                    ws.cell(row=row, column=6, value="⚠️ Превышено 24 часа")
                
                row += 1
        
        # Итоги по сотруднику (только если есть данные)
        if days_with_data > 0:
            ws.cell(row=row, column=2, value="ИТОГО:").font = Font(bold=True)
            ws.cell(row=row, column=3, value=total_day_hours).font = Font(bold=True)
            ws.cell(row=row, column=4, value=total_night_hours).font = Font(bold=True)
            ws.cell(row=row, column=5, value=total_day_hours + total_night_hours).font = Font(bold=True)
        else:
            ws.cell(row=row, column=2, value="ИТОГО:").font = Font(bold=True)
            ws.cell(row=row, column=3, value="-").font = Font(bold=True)
            ws.cell(row=row, column=4, value="-").font = Font(bold=True)
            ws.cell(row=row, column=5, value="-").font = Font(bold=True)
        
        row += 2
    
    # Добавляем авторскую подпись
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="Разработано и создано А.П. Плеханов")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=1).font = Font(italic=True, color="808080")
    
    # Настройка ширины колонок
    column_widths = [25, 15, 15, 15, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Сохраняем в байтовый поток
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def get_employee_time_data(period_start=None, period_end=None, specific_employee_id=None):
    """Получает данные о времени работы сотрудников за период"""
    # Получаем всех сотрудников
    employees = employees_sheet.get_all_records()
    employees_dict = {emp['ID']: emp['ФИО'] for emp in employees if emp.get('Активен', 'Да') == 'Да'}
    
    # Получаем записи о времени
    time_entries = time_sheet.get_all_records()
    
    # Структура для хранения данных: {ФИО: {дата: {'день': часы, 'ночь': часы}}}
    employee_data = {}
    
    for entry in time_entries:
        employee_id = str(entry['ID_сотрудника'])
        
        # Если запрашиваем конкретного сотрудника, пропускаем остальных
        if specific_employee_id and employee_id != str(specific_employee_id):
            continue
            
        if employee_id in employees_dict:
            employee_name = employees_dict[employee_id]
            date = entry['Дата']
            hours = float(entry['Часы'])
            shift_type = entry['Тип']
            
            # Проверяем период
            if period_start and period_end:
                try:
                    entry_date = datetime.strptime(date, '%d.%m.%Y')
                    start_date = datetime.strptime(period_start, '%d.%m.%Y')
                    end_date = datetime.strptime(period_end, '%d.%m.%Y')
                    
                    if not (start_date <= entry_date <= end_date):
                        continue
                except ValueError:
                    # Если дата в неправильном формате, пропускаем
                    continue
            
            if employee_name not in employee_data:
                employee_data[employee_name] = {}
            
            if date not in employee_data[employee_name]:
                employee_data[employee_name][date] = {'день': 0, 'ночь': 0}
            
            employee_data[employee_name][date][shift_type] += hours
    
    return employee_data

def validate_period(period_start, period_end, max_days=180):
    """Проверяет корректность периода и его длительность"""
    try:
        start_date = datetime.strptime(period_start, '%d.%m.%Y')
        end_date = datetime.strptime(period_end, '%d.%m.%Y')
        
        if start_date > end_date:
            return False, "❌ Дата начала не может быть позже даты окончания"
        
        period_days = (end_date - start_date).days
        if period_days < 0:
            return False, "❌ Период не может быть отрицательным"
        
        if period_days > max_days:
            return False, f"❌ Период не может превышать {max_days} дней"
        
        return True, f"✅ Период корректен: {period_days + 1} дней"
    
    except ValueError:
        return False, "❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ"

def get_available_dates():
    """Получает список всех дат, для которых есть записи в базе"""
    try:
        time_entries = time_sheet.get_all_records()
        dates = set()
        for entry in time_entries:
            dates.add(entry['Дата'])
        return sorted(list(dates))
    except Exception as e:
        print(f"Ошибка при получении списка дат: {e}")
        return []

# ========== ОБРАБОТЧИКИ КОМАНД ==========

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    user_id = update.effective_user.id
    print(f"DEBUG: Команда /start от пользователя {user_id}")
    
    if is_manager(user_id):
        # Меню для менеджера
        keyboard = [
            [InlineKeyboardButton("📊 Отчет по всем", callback_data="report_all")],
            [InlineKeyboardButton("👤 Отчет по сотруднику", callback_data="report_employee")],
            [InlineKeyboardButton("➕ Добавить сотрудника", callback_data="add_employee")],
            [InlineKeyboardButton("➖ Удалить сотрудника", callback_data="remove_employee")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("👨‍💼 Вы менеджер. Выберите действие:", reply_markup=reply_markup)
        print(f"DEBUG: Отправлено меню менеджера пользователю {user_id}")
    else:
        # Проверяем, зарегистрирован ли уже сотрудник
        employees = employees_sheet.get_all_records()
        user_found = None
        for employee in employees:
            if str(employee.get('TelegramID', '')) == str(user_id):
                user_found = employee
                break
        
        if user_found:
            # Сотрудник уже зарегистрирован
            keyboard = [
                [InlineKeyboardButton("⏱ Внести время", callback_data="add_time")],
                [InlineKeyboardButton("📋 Мои записи (90 дней)", callback_data="my_entries_90")],
                [InlineKeyboardButton("📅 Запросить отчет", callback_data="request_personal_report")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text(f"👋 Добро пожаловать, {user_found['ФИО']}!", reply_markup=reply_markup)
        else:
            # Выбор сотрудника из списка
            keyboard = []
            employees = employees_sheet.get_all_records()
            for employee in employees:
                if employee.get('Активен', 'Да') == 'Да' and not employee.get('TelegramID'):
                    keyboard.append([InlineKeyboardButton(employee['ФИО'], callback_data=f"select_{employee['ID']}")])
            
            if keyboard:
                reply_markup = InlineKeyboardMarkup(keyboard)
                await update.message.reply_text("Выберите свое ФИО из списка:", reply_markup=reply_markup)
            else:
                await update.message.reply_text("❌ Нет доступных сотрудников для выбора. Обратитесь к менеджеру.")

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик нажатий на кнопки"""
    query = update.callback_query
    user_id = update.effective_user.id
    data = query.data
    
    print(f"DEBUG: Получен callback от пользователя {user_id}: {data}")
    
    # Всегда отвечаем на callback_query, даже если это игнорируемая кнопка
    await query.answer()
    
    # Игнорируем кнопки с callback_data="ignore"
    if data == "ignore":
        return
    
    try:
        if data.startswith("select_"):
            # Сотрудник выбирает себя из списка
            employee_id = data.split("_")[1]
            
            # Обновляем TelegramID сотрудника
            employees = employees_sheet.get_all_values()
            for i, row in enumerate(employees[1:], start=2):  # Пропускаем заголовок
                if row[0] == employee_id:  # ID в первом столбце
                    employees_sheet.update_cell(i, 3, user_id)  # TelegramID в третьем столбце
                    employee_name = row[1]  # ФИО во втором столбце
                    break
            
            keyboard = [
                [InlineKeyboardButton("⏱ Внести время", callback_data="add_time")],
                [InlineKeyboardButton("📋 Мои записи (90 дней)", callback_data="my_entries_90")],
                [InlineKeyboardButton("📅 Запросить отчет", callback_data="request_personal_report")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.edit_message_text(f"✅ Вы успешно выбрали: {employee_name}", reply_markup=reply_markup)
        
        elif data == "add_time":
            # Создаем календарь для выбора даты
            today = datetime.now()
            keyboard = create_calendar(today.year, today.month)
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.edit_message_text("Выберите дату:", reply_markup=reply_markup)
        
        elif data.startswith("calendar_"):
            # Обработка выбора даты из календаря
            parts = data.split("_")
            year, month, day = int(parts[1]), int(parts[2]), int(parts[3])
            selected_date = f"{day:02d}.{month:02d}.{year}"
            
            user_data[user_id] = {'date': selected_date}
            
            await query.edit_message_text(
                f"📅 Выбрана дата: {selected_date}\n"
                f"Введите количество отработанных часов (например: 8 или 7.5):"
            )
            context.user_data['waiting_for_hours'] = True
        
        elif data.startswith("nav_"):
            # Навигация по календарю
            parts = data.split("_")
            year, month = int(parts[1]), int(parts[2])
            keyboard = create_calendar(year, month)
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.edit_message_text("Выберите дату:", reply_markup=reply_markup)
        
        elif data in ["type_day", "type_night"]:
            # Обработка выбора типа смены
            await handle_shift_type(update, context)
        
        elif data == "my_entries_90":
            await show_my_entries_90_days(update, context)
        
        elif data == "request_personal_report":
            await request_personal_report_period(update, context)
        
        elif data.startswith("remove_"):
            await handle_employee_removal(update, context)
        
        elif data.startswith("report_"):
            if data == "report_all":
                await request_report_period(update, context, "all")
            elif data == "report_employee":
                await select_employee_for_report(update, context)
            else:
                # report_{employee_id}
                employee_id = data.split("_")[1]
                context.user_data['selected_employee_id'] = employee_id
                await request_report_period(update, context, "employee")
        
        elif data in ["report_all", "report_employee", "add_employee", "remove_employee"]:
            # Обработка команд менеджера
            await handle_manager_commands(update, context, data)
        
        else:
            print(f"DEBUG: Неизвестный callback: {data}")
            await query.edit_message_text("❌ Неизвестная команда")
    
    except Exception as e:
        print(f"ERROR: Ошибка в обработчике кнопок: {e}")
        await query.edit_message_text("❌ Произошла ошибка при обработке команды")

async def show_my_entries_90_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает записи сотрудника за последние 90 дней"""
    query = update.callback_query
    user_id = update.effective_user.id
    
    # Получаем ID сотрудника
    employees = employees_sheet.get_all_records()
    employee_id = None
    employee_name = None
    for employee in employees:
        if str(employee.get('TelegramID', '')) == str(user_id):
            employee_id = employee['ID']
            employee_name = employee['ФИО']
            break
    
    if not employee_id:
        await query.edit_message_text("❌ Ошибка: сотрудник не найден")
        return
    
    # Получаем записи за последние 90 дней
    end_date = datetime.now()
    start_date = end_date - timedelta(days=90)
    
    employee_data = get_employee_time_data(
        period_start=start_date.strftime('%d.%m.%Y'),
        period_end=end_date.strftime('%d.%m.%Y'),
        specific_employee_id=employee_id
    )
    
    if employee_name in employee_data and employee_data[employee_name]:
        message = f"📋 Ваши записи за последние 90 дней:\n\n"
        total_day = 0
        total_night = 0
        days_with_entries = 0
        
        for date, shifts in sorted(employee_data[employee_name].items()):
            day_hours = shifts.get('день', 0)
            night_hours = shifts.get('ночь', 0)
            
            if day_hours > 0 or night_hours > 0:
                total_day += day_hours
                total_night += night_hours
                days_with_entries += 1
                
                message += f"📅 {date}:\n"
                if day_hours > 0:
                    message += f"   🌞 День: {day_hours} ч.\n"
                else:
                    message += f"   🌞 День: -\n"
                if night_hours > 0:
                    message += f"   🌙 Ночь: {night_hours} ч.\n"
                else:
                    message += f"   🌙 Ночь: -\n"
                message += f"   📊 Всего: {day_hours + night_hours} ч.\n\n"
        
        message += f"📈 ИТОГО за 90 дней:\n"
        message += f"📅 Дней с записями: {days_with_entries}\n"
        message += f"🌞 Всего дневных часов: {total_day} ч.\n"
        message += f"🌙 Всего ночных часов: {total_night} ч.\n"
        message += f"📊 Общее время: {total_day + total_night} ч.\n"
        message += f"📊 Среднее в день: {round((total_day + total_night) / max(days_with_entries, 1), 1)} ч."
        
        # Разбиваем сообщение если оно слишком длинное
        if len(message) > 4000:
            parts = [message[i:i+4000] for i in range(0, len(message), 4000)]
            for part in parts:
                await query.edit_message_text(part)
                # Добавляем небольшую задержку между сообщениями
                time.sleep(0.5)
        else:
            await query.edit_message_text(message)
    else:
        await query.edit_message_text("📭 У вас нет записей за последние 90 дней.")

async def request_personal_report_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запрашивает период для персонального отчета сотрудника"""
    query = update.callback_query
    
    # Получаем доступные даты для подсказки
    available_dates = get_available_dates()
    hint = ""
    if available_dates:
        hint = f"\n\n💡 Подсказка: в базе есть данные с {available_dates[0]} по {available_dates[-1]}"
    
    await query.edit_message_text(
        "📅 Введите период для вашего отчета в формате:\n"
        "ДД.ММ.ГГГГ ДД.ММ.ГГГГ\n\n"
        "Например: 01.11.2023 30.11.2023\n\n"
        "Максимальный период - 180 дней (6 месяцев)." + hint
    )
    context.user_data['waiting_for_personal_report_period'] = True

async def generate_personal_report(update: Update, context: ContextTypes.DEFAULT_TYPE, period_start: str, period_end: str):
    """Генерирует персональный отчет для сотрудника"""
    user_id = update.effective_user.id
    
    try:
        # Получаем ID сотрудника
        employees = employees_sheet.get_all_records()
        employee_id = None
        employee_name = None
        for employee in employees:
            if str(employee.get('TelegramID', '')) == str(user_id):
                employee_id = employee['ID']
                employee_name = employee['ФИО']
                break
        
        if not employee_id:
            await update.message.reply_text("❌ Ошибка: сотрудник не найден")
            return
        
        # Проверяем период
        is_valid, message = validate_period(period_start, period_end, max_days=180)
        if not is_valid:
            await update.message.reply_text(message)
            return
        
        # Получаем данные для отчета
        employee_data = get_employee_time_data(period_start, period_end, employee_id)
        
        # Всегда генерируем отчет, даже если нет данных
        # Генерируем Excel файл
        excel_file = generate_excel_report(employee_data, period_start, period_end)
        
        # Определяем имя файла
        filename = f"Мой_отчет_{employee_name}_{period_start}_{period_end}.xlsx"
        caption = f"📊 Ваш отчет за период с {period_start} по {period_end}"
        
        # Если нет данных, добавляем предупреждение
        if not employee_data or employee_name not in employee_data:
            caption += "\n\n⚠️ Внимание: за указанный период нет записей. В отчете отображены все дни периода с прочерками."
        
        # Отправляем файл
        await update.message.reply_document(
            document=excel_file,
            filename=filename,
            caption=caption
        )
        
    except Exception as e:
        print(f"ERROR: Ошибка генерации персонального отчета: {e}")
        await update.message.reply_text("❌ Произошла ошибка при генерации отчета.")

async def handle_manager_commands(update: Update, context: ContextTypes.DEFAULT_TYPE, command: str):
    """Обработка команд менеджера"""
    query = update.callback_query
    
    if command == "report_all":
        await request_report_period(update, context, "all")
    
    elif command == "report_employee":
        await select_employee_for_report(update, context)
    
    elif command == "add_employee":
        await add_employee_start(update, context)
    
    elif command == "remove_employee":
        await remove_employee_start(update, context)

async def request_report_period(update: Update, context: ContextTypes.DEFAULT_TYPE, report_type: str):
    """Запрашивает период для отчета менеджера"""
    query = update.callback_query
    
    # Получаем доступные даты для подсказки
    available_dates = get_available_dates()
    hint = ""
    if available_dates:
        hint = f"\n\n💡 Подсказка: в базе есть данные с {available_dates[0]} по {available_dates[-1]}"
    
    context.user_data['report_type'] = report_type
    await query.edit_message_text(
        "📅 Введите период для отчета в формате:\n"
        "ДД.ММ.ГГГГ ДД.ММ.ГГГГ\n\n"
        "Например: 01.11.2023 30.11.2023\n\n"
        "Максимальный период - 180 дней (6 месяцев)." + hint
    )
    context.user_data['waiting_for_report_period'] = True

async def select_employee_for_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выбор сотрудника для отчета"""
    query = update.callback_query
    
    # Получаем список активных сотрудников
    employees = employees_sheet.get_all_records()
    keyboard = []
    
    for employee in employees:
        if employee.get('Активен', 'Да') == 'Да':
            keyboard.append([InlineKeyboardButton(
                employee['ФИО'], 
                callback_data=f"report_{employee['ID']}"
            )])
    
    if keyboard:
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Выберите сотрудника для отчета:", reply_markup=reply_markup)
    else:
        await query.edit_message_text("❌ Нет активных сотрудников для отчета.")

async def generate_and_send_report(update: Update, context: ContextTypes.DEFAULT_TYPE, period_start: str, period_end: str):
    """Генерирует и отправляет отчет менеджеру"""
    user_id = update.effective_user.id
    
    try:
        # Проверяем период
        is_valid, message = validate_period(period_start, period_end, max_days=180)
        if not is_valid:
            await update.message.reply_text(message)
            return
        
        report_type = context.user_data.get('report_type', 'all')
        specific_employee_id = context.user_data.get('selected_employee_id')
        
        # Получаем данные для отчета
        employee_data = get_employee_time_data(period_start, period_end, specific_employee_id)
        
        # Всегда генерируем отчет, даже если нет данных
        # Генерируем Excel файл
        excel_file = generate_excel_report(employee_data, period_start, period_end)
        
        # Определяем имя файла
        if report_type == 'all':
            filename = f"Отчет_все_сотрудники_{period_start}_{period_end}.xlsx"
            caption = f"📊 Отчет по всем сотрудникам за период с {period_start} по {period_end}"
        else:
            # Находим имя сотрудника
            employees = employees_sheet.get_all_records()
            employee_name = next((emp['ФИО'] for emp in employees if str(emp['ID']) == str(specific_employee_id)), "Сотрудник")
            filename = f"Отчет_{employee_name}_{period_start}_{period_end}.xlsx"
            caption = f"👤 Отчет по сотруднику {employee_name} за период с {period_start} по {period_end}"
        
        # Если нет данных, добавляем предупреждение
        if not employee_data:
            caption += "\n\n⚠️ Внимание: за указанный период нет записей. В отчете отображены все дни периода с прочерками."
        
        # Отправляем файл
        await update.message.reply_document(
            document=excel_file,
            filename=filename,
            caption=caption
        )
        
        # Очищаем временные данные
        if 'selected_employee_id' in context.user_data:
            del context.user_data['selected_employee_id']
        if 'report_type' in context.user_data:
            del context.user_data['report_type']
        
    except Exception as e:
        print(f"ERROR: Ошибка генерации отчета: {e}")
        await update.message.reply_text("❌ Произошла ошибка при генерации отчета.")

# ... остальной код без изменений (функции add_employee_start, handle_employee_name_input, remove_employee_start, handle_employee_removal, handle_hours_input, handle_shift_type, main)

async def add_employee_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало процесса добавления сотрудника"""
    query = update.callback_query
    await query.edit_message_text(
        "➕ Добавление нового сотрудника\n\n"
        "Введите ФИО сотрудника:"
    )
    context.user_data['waiting_for_employee_name'] = True

async def handle_employee_name_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ввода ФИО сотрудника"""
    user_id = update.effective_user.id
    
    if context.user_data.get('waiting_for_employee_name'):
        employee_name = update.message.text
        
        # Получаем всех сотрудников для определения следующего ID
        employees = employees_sheet.get_all_records()
        new_id = len(employees) + 1
        
        # Добавляем нового сотрудника
        employees_sheet.append_row([new_id, employee_name, '', 'Да'])
        
        await update.message.reply_text(f"✅ Сотрудник '{employee_name}' успешно добавлен!")
        
        # Возвращаемся в меню менеджера
        keyboard = [
            [InlineKeyboardButton("📊 Отчет по всем", callback_data="report_all")],
            [InlineKeyboardButton("👤 Отчет по сотруднику", callback_data="report_employee")],
            [InlineKeyboardButton("➕ Добавить сотрудника", callback_data="add_employee")],
            [InlineKeyboardButton("➖ Удалить сотрудника", callback_data="remove_employee")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("👨‍💼 Вы менеджер. Выберите действие:", reply_markup=reply_markup)
        
        context.user_data['waiting_for_employee_name'] = False

async def remove_employee_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало процесса удаления сотрудника"""
    query = update.callback_query
    
    # Получаем список активных сотрудников
    employees = employees_sheet.get_all_records()
    keyboard = []
    
    for employee in employees:
        if employee.get('Активен', 'Да') == 'Да':
            keyboard.append([InlineKeyboardButton(
                f"{employee['ФИО']} (ID: {employee['ID']})", 
                callback_data=f"remove_{employee['ID']}"
            )])
    
    if keyboard:
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Выберите сотрудника для удаления:", reply_markup=reply_markup)
    else:
        await query.edit_message_text("❌ Нет активных сотрудников для удаления.")

async def handle_employee_removal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик удаления сотрудника"""
    query = update.callback_query
    data = query.data
    employee_id = data.split("_")[1]
    
    # Находим сотрудника и помечаем как неактивного
    employees = employees_sheet.get_all_values()
    for i, row in enumerate(employees[1:], start=2):  # Пропускаем заголовок
        if row[0] == employee_id:
            employees_sheet.update_cell(i, 4, 'Нет')  # Статус "Активен" в четвертом столбце
            employee_name = row[1]
            break
    
    await query.edit_message_text(f"✅ Сотрудник '{employee_name}' удален (деактивирован).")
    
    # Возвращаемся в меню менеджера
    keyboard = [
        [InlineKeyboardButton("📊 Отчет по всем", callback_data="report_all")],
        [InlineKeyboardButton("👤 Отчет по сотруднику", callback_data="report_employee")],
        [InlineKeyboardButton("➕ Добавить сотрудника", callback_data="add_employee")],
        [InlineKeyboardButton("➖ Удалить сотрудника", callback_data="remove_employee")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.message.reply_text("👨‍💼 Вы менеджер. Выберите действие:", reply_markup=reply_markup)

async def handle_hours_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ввода часов и периодов"""
    user_id = update.effective_user.id
    
    if context.user_data.get('waiting_for_hours'):
        try:
            hours = float(update.message.text)
            if hours <= 0 or hours > 24:
                await update.message.reply_text("❌ Часы должны быть от 0 до 24. Попробуйте снова:")
                return
            
            selected_date = user_data[user_id]['date']
            user_data[user_id]['hours'] = hours
            
            keyboard = [
                [InlineKeyboardButton("🌞 День", callback_data="type_day")],
                [InlineKeyboardButton("🌙 Ночь", callback_data="type_night")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"⏱ Часы: {hours}\n"
                f"📅 Дата: {selected_date}\n"
                "Выберите тип смены:",
                reply_markup=reply_markup
            )
            context.user_data['waiting_for_hours'] = False
            
        except ValueError:
            await update.message.reply_text("❌ Введите число (например: 8 или 7.5):")
    
    elif context.user_data.get('waiting_for_employee_name'):
        await handle_employee_name_input(update, context)
    
    elif context.user_data.get('waiting_for_report_period'):
        try:
            period_input = update.message.text.strip()
            dates = period_input.split()
            
            if len(dates) != 2:
                await update.message.reply_text("❌ Неверный формат. Введите две даты через пробел.")
                return
            
            period_start, period_end = dates
            
            await update.message.reply_text("⏳ Генерирую отчет...")
            await generate_and_send_report(update, context, period_start, period_end)
            
            context.user_data['waiting_for_report_period'] = False
            
        except Exception as e:
            await update.message.reply_text(f"❌ Ошибка: {e}")
    
    elif context.user_data.get('waiting_for_personal_report_period'):
        try:
            period_input = update.message.text.strip()
            dates = period_input.split()
            
            if len(dates) != 2:
                await update.message.reply_text("❌ Неверный формат. Введите две даты через пробел.")
                return
            
            period_start, period_end = dates
            
            await update.message.reply_text("⏳ Генерирую ваш отчет...")
            await generate_personal_report(update, context, period_start, period_end)
            
            context.user_data['waiting_for_personal_report_period'] = False
            
        except Exception as e:
            await update.message.reply_text(f"❌ Ошибка: {e}")

async def handle_shift_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик выбора типа смены"""
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    shift_type = "день" if query.data == "type_day" else "ночь"
    
    if user_id in user_data:
        date = user_data[user_id]['date']
        hours = user_data[user_id]['hours']
        
        # Получаем ID сотрудника
        employees = employees_sheet.get_all_records()
        employee_id = None
        employee_name = None
        for employee in employees:
            if str(employee.get('TelegramID', '')) == str(user_id):
                employee_id = employee['ID']
                employee_name = employee['ФИО']
                break
        
        if employee_id:
            # Проверяем, есть ли уже запись на эту дату и тип
            time_entries = time_sheet.get_all_records()
            existing_entry = None
            
            for entry in time_entries:
                if (str(entry['ID_сотрудника']) == str(employee_id) and 
                    entry['Дата'] == date and 
                    entry['Тип'] == shift_type):
                    existing_entry = entry
                    break
            
            if existing_entry:
                # Обновляем существующую запись
                for i, row in enumerate(time_sheet.get_all_values()[1:], start=2):
                    if (row[1] == str(employee_id) and row[2] == date and row[4] == shift_type):
                        time_sheet.update_cell(i, 4, hours)  # Обновляем часы
                        time_sheet.update_cell(i, 6, datetime.now().strftime('%d.%m.%Y %H:%M:%S'))
                        break
                
                await query.edit_message_text(f"✅ Запись обновлена!\nДата: {date}\nЧасы: {hours}\nТип: {shift_type}")
            else:
                # Добавляем новую запись
                new_id = len(time_entries) + 1
                new_row = [new_id, employee_id, date, hours, shift_type, datetime.now().strftime('%d.%m.%Y %H:%M:%S')]
                time_sheet.append_row(new_row)
                
                await query.edit_message_text(f"✅ Время успешно добавлено!\nДата: {date}\nЧасы: {hours}\nТип: {shift_type}")
            
            # Очищаем временные данные
            if user_id in user_data:
                del user_data[user_id]
        else:
            await query.edit_message_text("❌ Ошибка: сотрудник не найден")

def main():
    """Основная функция запуска бота"""
    print("🔄 Инициализация бота...")
    
    # Инициализируем Google Sheets
    if not initialize_google_sheets():
        print("❌ Не удалось инициализировать Google Sheets. Проверьте настройки.")
        return
    
    # Создаем приложение бота
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Добавляем обработчики в правильном порядке
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(button_handler))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_hours_input))
    
    # Запускаем бота
    print("✅ Бот запущен и готов к работе!")
    print("📱 Перейдите в Telegram и начните общение с ботом")
    
    try:
        application.run_polling(
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=True
        )
    except Exception as e:
        print(f"❌ Ошибка при запуске бота: {e}")

if __name__ == "__main__":
    main()